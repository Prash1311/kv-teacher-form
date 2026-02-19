from flask import Flask, request, jsonify
from flask_cors import CORS
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
from datetime import datetime

print("KV GOOGLE SHEET SERVER STARTING")
app = Flask(__name__)
CORS(app)

SHEET_NAME = "KV_Teacher_Data"
CREDS_FILE  = "creds.json"

# ── All columns in exact order ──────────────────────────────────────────────
COLUMNS = [
    "Timestamp",
    "RegistrationNo",
    "Name",
    "FatherName",
    "Gender",
    "Category",
    "DateOfBirth",
    "Age",
    "Mobile",
    "Email",
    "Aadhar",
    "PAN",
    "Address",
    "PostApplied",
    "Subject",
    # Flat qualification columns (easy to filter in Sheets)
    "XII_Year",
    "XII_Pct",
    "XII_Board",
    "Grad_Name",
    "Grad_Year",
    "Grad_Pct",
    "Grad_University",
    "PG_Name",
    "PG_Year",
    "PG_Pct",
    "PG_University",
    "BEd_Name",
    "BEd_Year",
    "BEd_Pct",
    # Additional info
    "CTET",
    "CTETScore",
    "TotalExp",
    "Languages",
    "OtherInfo",
    # Declaration
    "DeclPlace",
    "DeclDate",
    # Full structured data as JSON strings
    "Qualifications",
    "Experience",
]

# ── Sheet connection ────────────────────────────────────────────────────────
def connect_sheet():
    try:
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds  = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, scope)
        client = gspread.authorize(creds)
        sheet  = client.open(SHEET_NAME).sheet1
        print("✅ Connected to Google Sheet:", SHEET_NAME)

        # Write header row if the sheet is empty
        if sheet.row_count == 0 or sheet.cell(1, 1).value != "Timestamp":
            sheet.insert_row(COLUMNS, index=1)
            print("📝 Header row written.")

        return sheet
    except Exception as e:
        print("❌ SHEET CONNECTION ERROR:", str(e))
        return None

sheet = connect_sheet()

# ── Routes ──────────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return "KV Teacher Google Sheet Server Running ✅"


@app.route("/submit", methods=["POST"])
def submit():
    global sheet
    try:
        if sheet is None:
            sheet = connect_sheet()
            if sheet is None:
                return jsonify({"status": "error", "message": "Cannot connect to Google Sheet"}), 500

        data = request.get_json(force=True)

        # Validation
        for field in ["Name", "Email", "Mobile"]:
            if not data.get(field):
                return jsonify({"status": "error", "message": f"{field} is required"}), 400

        # Unique registration number
        reg_no = "KV-KRP-" + datetime.now().strftime("%Y%m%d%H%M%S")

        # Build row in same order as COLUMNS
        row = [reg_no if col == "RegistrationNo"
               else datetime.now().strftime("%d-%m-%Y %H:%M:%S") if col == "Timestamp"
               else data.get(col, "")
               for col in COLUMNS]

        sheet.append_row(row, value_input_option="USER_ENTERED")
        print(f"✅ Saved: {reg_no} — {data.get('Name')}")

        return jsonify({"status": "saved", "registration": reg_no})

    except Exception as e:
        print("❌ WRITE FAILED:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/data", methods=["GET"])
def get_data():
    """Return all rows as a list of dicts for the dashboard."""
    global sheet
    try:
        if sheet is None:
            sheet = connect_sheet()
            if sheet is None:
                return jsonify([])

        records = sheet.get_all_records()   # uses row 1 as keys automatically
        return jsonify(records)

    except Exception as e:
        print("❌ READ FAILED:", str(e))
        return jsonify([]), 500


@app.route("/download", methods=["GET"])
def download_excel():
    """Stream all data as a .xlsx file."""
    global sheet
    try:
        import io
        import openpyxl
        from flask import send_file

        if sheet is None:
            sheet = connect_sheet()

        records = sheet.get_all_records()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        if records:
            # Header
            ws.append(list(records[0].keys()))
            # Data rows
            for rec in records:
                ws.append(list(rec.values()))

            # Basic formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="0D47A1")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"KV_Applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except ImportError:
        return jsonify({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}), 500
    except Exception as e:
        print("❌ DOWNLOAD FAILED:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Entry point ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)